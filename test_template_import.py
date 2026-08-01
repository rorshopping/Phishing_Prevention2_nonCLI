import asyncio, uuid, os, sys
sys.path.insert(0, "C:\\Users\\Richard\\Documents\\Projects\\Phishing_Prevention2_nonCLI")
os.environ["PYTHONPATH"] = "C:\\Users\\Richard\\Documents\\Projects\\Phishing_Prevention2_nonCLI"
os.chdir("C:\\Users\\Richard\\Documents\\Projects\\Phishing_Prevention2_nonCLI")

import logging
logging.basicConfig(level=logging.WARNING, format="%(asctime)s %(levelname)-8s %(name)s %(message)s")

from openpyxl import load_workbook
from datetime import datetime
from sqlalchemy import select

TEMPLATE_PATH = "client_onboarding_template.xlsx"

IMPORT_SHEET = "test_filled_template.xlsx"


def create_test_filled_xlsx():
    """Create a filled copy of the template with test data, exactly as a client would fill it."""
    wb = load_workbook(TEMPLATE_PATH)

    ws1 = wb["Client"]
    ws1.cell(row=5, column=1, value="TestClient AG")
    ws1.cell(row=5, column=2, value="admin@testclient.ag")
    ws1.cell(row=5, column=3, value="Hans Mueller")
    ws1.cell(row=5, column=4, value="Financial Services")
    ws1.cell(row=5, column=5, value=350)
    ws1.cell(row=5, column=6, value="DE")
    ws1.cell(row=5, column=7, value=6)
    ws1.cell(row=5, column=8, value="TRUE")

    ws2 = wb["Employees"]
    test_emps = [
        ["ceo@testclient.ag", "Klaus Schmidt", "CEO", "Executive", "executive", ""],
        ["cfo@testclient.ag", "Anna Weber", "CFO", "Finance", "finance", "https://linkedin.com/in/anna"],
        ["cto@testclient.ag", "Lukas Fischer", "CTO", "IT", "it_management", ""],
        ["helpdesk@testclient.ag", "Marie Braun", "Helpdesk", "IT", "it_staff", ""],
        ["hr@testclient.ag", "Sarah Klein", "HR Lead", "HR", "hr", ""],
        ["eng-lead@testclient.ag", "Tom Wagner", "Engineering Lead", "Engineering", "engineering",
         "https://linkedin.com/in/tom"],
        ["sales@testclient.ag", "Laura Hoffmann", "Sales Rep", "Sales", "sales", ""],
        ["staff@testclient.ag", "Julia Meyer", "Staff", "General", "general", ""],
    ]
    for i, emp in enumerate(test_emps):
        r = 8 + i
        for c, val in enumerate(emp, 1):
            ws2.cell(row=r, column=c, value=val)

    wb.save(IMPORT_SHEET)
    print(f"Created test data: {IMPORT_SHEET}")


async def import_and_verify():
    from src.database.session import async_session, engine, Base
    from src.database.models import Client, Employee, EmployeeGroup
    from src.agents.execution_agent import ExecutionAgent
    from src.database.models import Campaign, CampaignStatus

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    wb = load_workbook(IMPORT_SHEET)

    ws1 = wb["Client"]
    headers1 = [c.value.split("\n")[0] for c in ws1[3] if c.value]
    client_row = [c.value for c in ws1[5] if c.value is not None]
    client_data = dict(zip(headers1, client_row))

    print("\n=== PARSED CLIENT DATA ===")
    for k, v in client_data.items():
        print(f"  {k}: {v!r} ({type(v).__name__})")

    ws2 = wb["Employees"]
    headers2 = [c.value.split("\n")[0] for c in ws2[3] if c.value]
    employees_raw = []
    for row in ws2.iter_rows(min_row=8, max_row=15, values_only=True):
        if row[0]:
            non_none = {k: v for k, v in zip(headers2, row) if v is not None}
            employees_raw.append(non_none)

    print(f"\n=== PARSED EMPLOYEES ({len(employees_raw)}) ===")
    for e in employees_raw:
        print(f"  {e['email']:30s} group={e.get('group',''):15s} name={e.get('name','')}")

    print("\n--- Type validation ---")
    type_ok = True
    try:
        emp_count = int(client_data["employee_count"])
        print(f"  employee_count: {emp_count} OK")
    except (ValueError, TypeError) as e:
        print(f"  employee_count ERROR: {e}")
        type_ok = False

    valid_groups = {g.value for g in EmployeeGroup}
    for e in employees_raw:
        g = e.get("group", "general")
        if g and g not in valid_groups:
            print(f"  INVALID group for {e['email']}: '{g}' (valid: {valid_groups})")
            type_ok = False

    if not type_ok:
        print("\n[RED]Type validation FAILED — fix template data[/]")
        return

    print("\n--- Importing into database ---")
    async with async_session() as db:
        client = Client(
            company_name=client_data["company_name"],
            contact_email=client_data["contact_email"],
            contact_name=client_data.get("contact_name"),
            industry=client_data.get("industry"),
            employee_count=int(client_data["employee_count"]),
            country=client_data.get("country", "DE"),
            campaigns_per_year=int(client_data.get("campaigns_per_year", 25)),
            vishing_enabled=str(client_data.get("vishing_enabled", "FALSE")).upper() == "TRUE",
        )
        db.add(client)
        await db.flush()
        print(f"  Created client: {client.id} ({client.company_name})")

        from src.utils.gdpr import hash_pii

        created_employees = []
        for e in employees_raw:
            email = e.get("email", "")
            emp = Employee(
                client_id=client.id,
                email=email,
                email_hash=hash_pii(email),
                name=e.get("name"),
                role=e.get("role"),
                department=e.get("department"),
                group=e.get("group") or "general",
                linkedin_url=e.get("linkedin_url") or None,
            )
            db.add(emp)
            created_employees.append(emp)
        await db.flush()

        for emp in created_employees:
            await db.refresh(emp)
            print(f"  Created employee: {emp.id} {emp.email} ({emp.group.value})")

        campaign = Campaign(
            client_id=client.id,
            name=f"Onboarding Test {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}",
            status=CampaignStatus.draft,
            difficulty="easy",
        )
        db.add(campaign)
        await db.commit()
        await db.refresh(campaign)

        print(f"\n  Campaign: {campaign.id}")

    plan = {
        "client_id": str(client.id),
        "name": f"TestClient AG - {datetime.utcnow().strftime('%Y-%m-%d')}",
        "difficulty": "easy",
        "monthly_budget": 1,
        "industry_context": {"industry": "Financial Services", "country": "DE",
                            "employee_count": 350, "recent_threats": []},
        "llm_strategy": {"difficulty": "easy", "scenario_weights": {}, "rationale": "test"},
        "employee_assignments": [
            {"employee_id": str(created_employees[0].id), "group": "executive", "scenario_type": "ceo_fraud"},
            {"employee_id": str(created_employees[1].id), "group": "finance", "scenario_type": "invoice_fraud"},
            {"employee_id": str(created_employees[2].id), "group": "it_management", "scenario_type": "cloud_notification"},
            {"employee_id": str(created_employees[3].id), "group": "it_staff", "scenario_type": "malware_attachment"},
            {"employee_id": str(created_employees[4].id), "group": "hr", "scenario_type": "credential_harvest"},
            {"employee_id": str(created_employees[5].id), "group": "engineering", "scenario_type": "voicemail_phish"},
            {"employee_id": str(created_employees[6].id), "group": "sales", "scenario_type": "linkedin_message"},
            {"employee_id": str(created_employees[7].id), "group": "general", "scenario_type": "urgency_alert"},
        ],
    }

    print("\n--- Running ExecutionAgent with 8 scenarios ---")
    agent = ExecutionAgent()
    success = await agent.execute_campaign(campaign.id, plan)
    print(f"  Execution result: {success}")

    async with async_session() as db:
        result = await db.execute(select(Campaign).where(Campaign.id == campaign.id))
        updated = result.scalar_one()
        print(f"\n  Campaign status: {updated.status.value}")
        print(f"  Gophish IDs: {updated.gophish_campaign_id}")
        print(f"  Gophish group IDs: {updated.gophish_group_id}")
        print(f"  Gophish template IDs: {updated.gophish_template_id}")

        gophish_ids = updated.gophish_campaign_id.split(",") if updated.gophish_campaign_id else []
        print(f"\n  Gophish campaigns created: {len(gophish_ids)} (expected: 8)")

    print("\n=== IMPORT & EXECUTION COMPLETE ===")


async def main():
    create_test_filled_xlsx()
    await import_and_verify()


asyncio.run(main())
